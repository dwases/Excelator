from openpyxl import load_workbook
from openpyxl import Workbook
import re

new_wb = Workbook()  # nowy plik z wyodrebnionymi danymi
#old_wb = load_workbook("C:\\Users\\Antonio\\Desktop\\pacjenci BU.xlsx")
old_wb = load_workbook("pacjenci BU.xlsx")
# print(old_wb.sheetnames)
# czytanie ze starego pliku dziala
new_ws1 = new_wb.active  # wb.create_sheet("main sheet")  # work sheet
new_ws1.title = "sheet 1"
# new_ws1['F5'] = 3.14  # zapisywanie do komorki dziala
# print(new_ws1['F5'].value)

# adresy - kolumna C - old_wb - pierwsza kom. 'adres', reszta adresy
# 7469 - ostatni rekord
old_ws1 = old_wb.active  # czytanie ze starego pliku dziala
# print(old_ws1['C2'].value)

colA = old_ws1['A']
colB = old_ws1['B']
colC = old_ws1['C']
colG = old_ws1['G']

new_colA = new_ws1['A']
new_colB = new_ws1['B']
new_colC = new_ws1['C']
new_colD = new_ws1['D']
new_colE = new_ws1['E']
new_colF = new_ws1['F']


new_ws1['A1'] = "Nieskategoryzowane adresy"
new_ws1['B1'] = "Kody pocztowe"
new_ws1['C1'] = "Miasto i/lub Ulica"
new_ws1['D1'] = "Domy"
new_ws1['E1'] = "Mieszkania"
new_ws1['F1'] = "Skategoryzowane"
new_ws1['G1'] = "Nieskategoryzowane tnm"
new_ws1['H1'] = "Skategoryzowane tnm"
new_ws1['I1'] = "cT"
new_ws1['J1'] = "cN"
new_ws1['K1'] = "M"
new_ws1['L1'] = "pT"
new_ws1['M1'] = "pN"


for i in range(0, 7469):  # iterowanie po calej kolumnie C starego pliku
    # if old_ws1['C{}'.format(i)] == "adres":
    try:
        if colC[i].value == 'adres':
            print("adres")
            pass
        elif bool(re.match("[0-9][0-9]-[0-9][0-9][0-9].*", str(colC[i].value))):
            # print("skategoryzowano")
            new_ws1['F{}'.format(i+1)].value = colC[i].value
            new_ws1['B{}'.format(i+1)].value = colC[i].value[0:6]
            # do teraz dziala poprawnie
            if '/' in colC[i].value:  # mieszkania
                new_ws1['E{}'.format(i+1)].value = colC[i].value.split(" ")[-1]
            else:  # domy
                new_ws1['D{}'.format(i+1)].value = colC[i].value.split(" ")[-1]
            # kolumna miasto i/lub ulica
            new_ws1['C{}'.format(i+1)].value = ' '.join(colC[i].value.split(" ")[1:-1])
        else:
            new_colA[i].value = colC[i].value
    except:
        new_ws1['A{}'.format(i+1)].value = colC[i].value
        # print(new_ws1['A{}'.format(i+1)].value)
        # print("nieskategoryzowane")


# teraz pora na tnm
for i in range(0, 7469):
    try:
        if colG[i].value == 'tnm':
            print("tnm")
            pass
        elif bool(re.match("[y]?[cCpP][T].*", str(colG[i].value))):
            # skategoryzowane
            new_ws1['H{}'.format(i+1)].value = colG[i].value
            # do teraz dziala
            # kolumna I czyli pT
            if bool(re.match("[y]?[pP][T].*", str(colG[i].value))):
                new_ws1['L{}'.format(
                    i+1)].value = colG[i].value[colG[i].value.index('T')+1:colG[i].value.index('N')]
                # kolumna K czyli pN
                new_ws1['M{}'.format(
                    i+1)].value = colG[i].value[colG[i].value.index('N')+1:colG[i].value.index('M')]

            # kolumna J czyli cT
            elif bool(re.match("[y]?[cC][T].*", str(colG[i].value))):
                new_ws1['I{}'.format(
                    i+1)].value = colG[i].value[colG[i].value.index('T')+1:colG[i].value.index('N')]
                # kolumna L czyli cN
                new_ws1['J{}'.format(
                    i+1)].value = colG[i].value[colG[i].value.index('N')+1:colG[i].value.index('M')]
            # kolumna L czyli M xD
            new_ws1['K{}'.format(
                i+1)].value = colG[i].value[colG[i].value.index('M')+1:]

        else:
            # nieskategoryzowane
            new_ws1['G{}'.format(i+1)].value = colG[i].value
    except:
        new_ws1['G{}'.format(i+1)].value = colG[i].value
new_wb.save(filename='pacjenci BU dane.xlsx')
